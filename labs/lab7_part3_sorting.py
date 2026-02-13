#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Лабораторная работа №7 (Часть 3)
Microsoft Office Excel. Сортировка, фильтры и промежуточные итоги
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import random

def create_sorting_filtering():
    """Создание таблицы с сортировкой и фильтрами"""
    
    # Создание новой книги
    wb = openpyxl.Workbook()
    
    # Лист 1: Сортировка
    ws_sort = wb.active
    ws_sort.title = "Сортировка"
    
    # Заголовок
    ws_sort['A1'] = "УЧЕТ ТОВАРОВ НА СКЛАДЕ"
    ws_sort['A1'].font = Font(bold=True, size=14)
    ws_sort.merge_cells('A1:G1')
    ws_sort['A1'].alignment = Alignment(horizontal='center')
    
    # Заголовки таблицы
    headers = ['№', 'Код товара', 'Наименование товара', 'Дата поступления', 
               'Количество', 'Цена (руб.)', 'Стоимость (руб.)']
    
    for col, header in enumerate(headers, 1):
        cell = ws_sort.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    # Данные товаров
    товары = [
        "Компьютер",
        "Ноутбук",
        "Монитор",
        "Клавиатура",
        "Мышь"
    ]
    
    # Генерация данных
    data = []
    start_date = datetime(2024, 1, 1)
    
    for i in range(15):
        товар = random.choice(товары)
        код = f"T{random.randint(1000, 9999)}"
        дата = start_date + timedelta(days=random.randint(0, 100))
        количество = random.randint(5, 50)
        
        # Цены в зависимости от товара
        цены = {
            "Компьютер": random.randint(45000, 65000),
            "Ноутбук": random.randint(35000, 55000),
            "Монитор": random.randint(10000, 25000),
            "Клавиатура": random.randint(800, 2500),
            "Мышь": random.randint(300, 1500)
        }
        цена = цены[товар]
        
        data.append([i+1, код, товар, дата, количество, цена])
    
    # Сортировка по наименованию, затем по дате
    data.sort(key=lambda x: (x[2], x[3]))
    
    # Заполнение таблицы
    for row_idx, row_data in enumerate(data, 4):
        ws_sort[f'A{row_idx}'] = row_data[0]  # №
        ws_sort[f'B{row_idx}'] = row_data[1]  # Код
        ws_sort[f'C{row_idx}'] = row_data[2]  # Наименование
        ws_sort[f'D{row_idx}'] = row_data[3].strftime('%d.%m.%Y')  # Дата
        ws_sort[f'E{row_idx}'] = row_data[4]  # Количество
        ws_sort[f'F{row_idx}'] = row_data[5]  # Цена
        ws_sort[f'F{row_idx}'].number_format = '#,##0.00 ₽'
        
        # Стоимость = Количество * Цена
        ws_sort[f'G{row_idx}'] = f'=E{row_idx}*F{row_idx}'
        ws_sort[f'G{row_idx}'].number_format = '#,##0.00 ₽'
    
    # Настройка ширины столбцов
    ws_sort.column_dimensions['A'].width = 5
    ws_sort.column_dimensions['B'].width = 12
    ws_sort.column_dimensions['C'].width = 20
    ws_sort.column_dimensions['D'].width = 15
    ws_sort.column_dimensions['E'].width = 12
    ws_sort.column_dimensions['F'].width = 15
    ws_sort.column_dimensions['G'].width = 18
    
    # Лист 2: Промежуточные итоги
    ws_itogi = wb.create_sheet("Итоги")
    
    # Копируем данные
    for row in ws_sort.iter_rows(min_row=1, max_row=len(data)+3):
        for cell in row:
            new_cell = ws_itogi[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
    
    # Копируем ширину столбцов
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_itogi.column_dimensions[col].width = ws_sort.column_dimensions[col].width
    
    # Добавление промежуточных итогов по товарам
    current_product = None
    subtotal_rows = []
    last_row = len(data) + 3
    insert_offset = 0
    
    for idx, row_data in enumerate(data, 4):
        actual_row = idx + insert_offset
        product = row_data[2]
        
        if current_product is not None and current_product != product:
            # Вставляем строку подытога
            ws_itogi.insert_rows(actual_row)
            ws_itogi[f'A{actual_row}'].value = f"Итого по {current_product}:"
            ws_itogi[f'A{actual_row}'].font = Font(bold=True, italic=True)
            ws_itogi.merge_cells(f'A{actual_row}:F{actual_row}')
            ws_itogi[f'A{actual_row}'].alignment = Alignment(horizontal='right')
            
            # Формула суммы
            start = subtotal_rows[0] if subtotal_rows else actual_row - 1
            end = actual_row - 1
            ws_itogi[f'G{actual_row}'] = f'=SUBTOTAL(9, G{start}:G{end})'
            ws_itogi[f'G{actual_row}'].number_format = '#,##0.00 ₽'
            ws_itogi[f'G{actual_row}'].font = Font(bold=True, italic=True)
            
            subtotal_rows = []
            insert_offset += 1
        
        current_product = product
        subtotal_rows.append(actual_row)
    
    # Последний подытог
    if current_product:
        final_row = last_row + insert_offset + 1
        ws_itogi.insert_rows(final_row)
        ws_itogi[f'A{final_row}'].value = f"Итого по {current_product}:"
        ws_itogi[f'A{final_row}'].font = Font(bold=True, italic=True)
        ws_itogi.merge_cells(f'A{final_row}:F{final_row}')
        ws_itogi[f'A{final_row}'].alignment = Alignment(horizontal='right')
        
        start = subtotal_rows[0]
        end = final_row - 1
        ws_itogi[f'G{final_row}'] = f'=SUBTOTAL(9, G{start}:G{end})'
        ws_itogi[f'G{final_row}'].number_format = '#,##0.00 ₽'
        ws_itogi[f'G{final_row}'].font = Font(bold=True, italic=True)
    
    # Общий итог
    grand_total_row = final_row + 1
    ws_itogi[f'A{grand_total_row}'].value = "ОБЩИЙ ИТОГ:"
    ws_itogi[f'A{grand_total_row}'].font = Font(bold=True, size=12)
    ws_itogi.merge_cells(f'A{grand_total_row}:F{grand_total_row}')
    ws_itogi[f'A{grand_total_row}'].alignment = Alignment(horizontal='right')
    
    ws_itogi[f'G{grand_total_row}'] = f'=SUBTOTAL(9, G4:G{grand_total_row-1})'
    ws_itogi[f'G{grand_total_row}'].number_format = '#,##0.00 ₽'
    ws_itogi[f'G{grand_total_row}'].font = Font(bold=True, size=12)
    ws_itogi[f'G{grand_total_row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Лист 3: Фильтр
    ws_filter = wb.create_sheet("Фильтр")
    
    # Копируем исходные данные
    for row in ws_sort.iter_rows(min_row=1, max_row=len(data)+3):
        for cell in row:
            new_cell = ws_filter[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
    
    # Копируем ширину столбцов
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_filter.column_dimensions[col].width = ws_sort.column_dimensions[col].width
    
    # Включаем автофильтр
    ws_filter.auto_filter.ref = f'A3:G{len(data)+3}'
    
    # Сохранение файла
    filename = 'labs/Lab7_Part3_Sorting.xlsx'
    wb.save(filename)
    print(f"✓ Лабораторная работа №7 (Часть 3) успешно создана: {filename}")
    print(f"  Создано 3 листа:")
    print(f"    - Сортировка (данные отсортированы)")
    print(f"    - Итоги (с промежуточными итогами)")
    print(f"    - Фильтр (с автофильтром)")
    print(f"  Всего записей: {len(data)}")
    return filename

if __name__ == "__main__":
    create_sorting_filtering()
