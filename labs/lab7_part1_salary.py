#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Лабораторная работа №7 (Часть 1)
Microsoft Office Excel. Формулы, функции и диаграммы.
Расчет заработной платы
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

def create_salary_table():
    """Создание таблицы расчета заработной платы"""
    
    # Создание новой книги
    wb = openpyxl.Workbook()
    
    # Лист со списком сотрудников (скрытый)
    ws_employees = wb.active
    ws_employees.title = "Сотрудники"
    
    employees = [
        "Иванов Иван Иванович",
        "Петров Петр Петрович",
        "Сидорова Анна Сергеевна",
        "Кузнецов Сергей Александрович",
        "Васильева Мария Дмитриевна",
        "Смирнов Алексей Викторович",
        "Новикова Елена Павловна"
    ]
    
    ws_employees['A1'] = "Список сотрудников"
    ws_employees['A1'].font = Font(bold=True, size=12)
    
    for i, emp in enumerate(employees, 2):
        ws_employees[f'A{i}'] = emp
    
    # Основной лист с расчетом
    ws = wb.create_sheet("Расчет заработной платы", 0)
    
    # Заголовок
    ws.merge_cells('A1:J1')
    ws['A1'] = "РАСЧЕТ ЗАРАБОТНОЙ ПЛАТЫ СОТРУДНИКОВ ПРЕДПРИЯТИЯ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Курс доллара
    ws['C3'] = "Курс доллара:"
    ws['D3'] = 32
    ws['D3'].font = Font(bold=True)
    
    # Заголовки таблицы
    headers = ['№', 'Ф.И.О.', 'Оклад (руб.)', 'Премия (руб.)', 
               'Итого начислено (руб.)', 'Подоходный налог (руб.)', 
               'Сумма к выдаче (руб.)', 'Сумма к выдаче ($)']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Данные сотрудников
    salaries = [25000, 30000, 28000, 35000, 27000, 32000, 29000]
    
    for i, (emp, salary) in enumerate(zip(employees, salaries), 6):
        row = i
        ws[f'A{row}'] = i - 5  # Номер
        ws[f'B{row}'] = emp  # ФИО
        ws[f'C{row}'] = salary  # Оклад
        ws[f'C{row}'].number_format = '#,##0.00 ₽'
        
        # Премия 20%
        ws[f'D{row}'] = f'=C{row}*0.2'
        ws[f'D{row}'].number_format = '#,##0.00 ₽'
        
        # Итого начислено
        ws[f'E{row}'] = f'=C{row}+D{row}'
        ws[f'E{row}'].number_format = '#,##0.00 ₽'
        
        # Подоходный налог 13%
        ws[f'F{row}'] = f'=E{row}*0.13'
        ws[f'F{row}'].number_format = '#,##0.00 ₽'
        
        # Сумма к выдаче
        ws[f'G{row}'] = f'=E{row}-F{row}'
        ws[f'G{row}'].number_format = '#,##0.00 ₽'
        
        # Сумма к выдаче в долларах
        ws[f'H{row}'] = f'=G{row}/$D$3'
        ws[f'H{row}'].number_format = '$#,##0.00'
    
    # Итоговая строка
    last_row = len(employees) + 6
    ws[f'A{last_row}'] = "ИТОГО:"
    ws[f'A{last_row}'].font = Font(bold=True)
    ws.merge_cells(f'A{last_row}:E{last_row}')
    ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
    
    # Сумма подоходного налога
    ws[f'F{last_row}'] = f'=SUM(F6:F{last_row-1})'
    ws[f'F{last_row}'].font = Font(bold=True)
    ws[f'F{last_row}'].number_format = '#,##0.00 ₽'
    
    # Общая сумма к выдаче в рублях
    ws[f'G{last_row}'] = f'=SUM(G6:G{last_row-1})'
    ws[f'G{last_row}'].font = Font(bold=True)
    ws[f'G{last_row}'].number_format = '#,##0.00 ₽'
    
    # Общая сумма к выдаче в долларах
    ws[f'H{last_row}'] = f'=SUM(H6:H{last_row-1})'
    ws[f'H{last_row}'].font = Font(bold=True)
    ws[f'H{last_row}'].number_format = '$#,##0.00'
    
    # Статистика
    stat_row = last_row + 2
    ws[f'A{stat_row}'] = "Средняя зарплата:"
    ws[f'B{stat_row}'] = f'=AVERAGE(G6:G{last_row-1})'
    ws[f'B{stat_row}'].number_format = '#,##0.00 ₽'
    
    ws[f'A{stat_row+1}'] = "Минимальная зарплата:"
    ws[f'B{stat_row+1}'] = f'=MIN(G6:G{last_row-1})'
    ws[f'B{stat_row+1}'].number_format = '#,##0.00 ₽'
    
    ws[f'A{stat_row+2}'] = "Максимальная зарплата:"
    ws[f'B{stat_row+2}'] = f'=MAX(G6:G{last_row-1})'
    ws[f'B{stat_row+2}'].number_format = '#,##0.00 ₽'
    
    # Условное форматирование - красный цвет для зарплат < 5500
    from openpyxl.styles import Color
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    # Диаграмма - Зарплата сотрудников
    chart = BarChart()
    chart.title = "Заработная плата сотрудников предприятия"
    chart.x_axis.title = "Сотрудники"
    chart.y_axis.title = "Сумма к выдаче (руб.)"
    
    # Данные для диаграммы
    data = Reference(ws, min_col=7, min_row=5, max_row=last_row-1)
    cats = Reference(ws, min_col=2, min_row=6, max_row=last_row-1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws.add_chart(chart, f'A{stat_row+5}')
    
    # Круговая диаграмма - соотношение налогов и выплат
    pie_chart = PieChart()
    pie_chart.title = "Структура выплат"
    
    # Создаем данные для круговой диаграммы
    pie_row = stat_row + 5
    ws[f'E{pie_row}'] = "Категория"
    ws[f'F{pie_row}'] = "Сумма"
    ws[f'E{pie_row+1}'] = "К выдаче"
    ws[f'F{pie_row+1}'] = f'=G{last_row}'
    ws[f'E{pie_row+2}'] = "Налоги"
    ws[f'F{pie_row+2}'] = f'=F{last_row}'
    
    labels = Reference(ws, min_col=5, min_row=pie_row+1, max_row=pie_row+2)
    data = Reference(ws, min_col=6, min_row=pie_row, max_row=pie_row+2)
    
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.height = 10
    pie_chart.width = 12
    
    ws.add_chart(pie_chart, f'K{stat_row+5}')
    
    # Сохранение файла
    filename = 'labs/Lab7_Part1_Salary.xlsx'
    wb.save(filename)
    print(f"✓ Лабораторная работа №7 (Часть 1) успешно создана: {filename}")
    print(f"  Обработано сотрудников: {len(employees)}")
    return filename

if __name__ == "__main__":
    create_salary_table()
