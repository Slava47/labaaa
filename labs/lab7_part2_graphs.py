#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Лабораторная работа №7 (Часть 2)
Microsoft Office Excel. Построение графиков функций
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import ScatterChart, Reference, Series
import math

def create_function_graphs():
    """Создание графиков функций"""
    
    # Создание новой книги
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Графики функций"
    
    # Заголовок
    ws.merge_cells('A1:F1')
    ws['A1'] = "ПОСТРОЕНИЕ ГРАФИКОВ ФУНКЦИЙ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # График 1: y = sin(x) на отрезке [-6; 6] с шагом 0.5
    ws['A3'] = "График 1: y = sin(x)"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A4'] = "x"
    ws['B4'] = "y = sin(x)"
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    
    # Генерация данных для sin(x)
    x_start = -6
    x_end = 6
    step = 0.5
    row = 5
    
    x = x_start
    while x <= x_end:
        ws[f'A{row}'] = x
        ws[f'B{row}'] = f'=SIN(A{row})'
        x += step
        row += 1
    
    # Создание графика sin(x)
    chart1 = ScatterChart()
    chart1.title = "График функции y = sin(x)"
    chart1.x_axis.title = "x"
    chart1.y_axis.title = "y"
    chart1.style = 13
    
    xvalues = Reference(ws, min_col=1, min_row=5, max_row=row-1)
    yvalues = Reference(ws, min_col=2, min_row=4, max_row=row-1)
    
    series = Series(yvalues, xvalues, title_from_data=True)
    chart1.series.append(series)
    chart1.height = 10
    chart1.width = 15
    
    ws.add_chart(chart1, 'D4')
    
    # График 2: кусочная функция на отрезке [-3; 3] с шагом 0.2
    start_row2 = row + 3
    ws[f'A{start_row2}'] = "График 2: Кусочная функция"
    ws[f'A{start_row2}'].font = Font(bold=True, size=12)
    
    ws[f'A{start_row2+1}'] = "x"
    ws[f'B{start_row2+1}'] = "y"
    ws[f'A{start_row2+1}'].font = Font(bold=True)
    ws[f'B{start_row2+1}'].font = Font(bold=True)
    
    # y = { 1 - x^2, если x ∈ [-1, 1]
    #     { |x| - 1, иначе
    
    x_start2 = -3
    x_end2 = 3
    step2 = 0.2
    row2 = start_row2 + 2
    
    x = x_start2
    while x <= x_end2 + 0.01:  # +0.01 для компенсации ошибок округления
        ws[f'A{row2}'] = round(x, 2)
        # Формула: =IF(AND(A{row2}>=-1, A{row2}<=1), 1-A{row2}*A{row2}, ABS(A{row2})-1)
        ws[f'B{row2}'] = f'=IF(AND(A{row2}>=-1, A{row2}<=1), 1-A{row2}*A{row2}, ABS(A{row2})-1)'
        x += step2
        row2 += 1
    
    # Создание графика для кусочной функции
    chart2 = ScatterChart()
    chart2.title = "График кусочной функции"
    chart2.x_axis.title = "x"
    chart2.y_axis.title = "y"
    chart2.style = 13
    
    xvalues2 = Reference(ws, min_col=1, min_row=start_row2+2, max_row=row2-1)
    yvalues2 = Reference(ws, min_col=2, min_row=start_row2+1, max_row=row2-1)
    
    series2 = Series(yvalues2, xvalues2, title_from_data=True)
    chart2.series.append(series2)
    chart2.height = 10
    chart2.width = 15
    
    ws.add_chart(chart2, f'D{start_row2}')
    
    # График 3: y = x^2 - 4
    start_row3 = row2 + 3
    ws[f'A{start_row3}'] = "График 3: y = x² - 4"
    ws[f'A{start_row3}'].font = Font(bold=True, size=12)
    
    ws[f'A{start_row3+1}'] = "x"
    ws[f'B{start_row3+1}'] = "y = x² - 4"
    ws[f'A{start_row3+1}'].font = Font(bold=True)
    ws[f'B{start_row3+1}'].font = Font(bold=True)
    
    # Данные для параболы
    x_start3 = -5
    x_end3 = 5
    step3 = 0.5
    row3 = start_row3 + 2
    
    x = x_start3
    while x <= x_end3:
        ws[f'A{row3}'] = x
        ws[f'B{row3}'] = f'=A{row3}*A{row3}-4'
        x += step3
        row3 += 1
    
    # Создание графика параболы
    chart3 = ScatterChart()
    chart3.title = "График функции y = x² - 4"
    chart3.x_axis.title = "x"
    chart3.y_axis.title = "y"
    chart3.style = 13
    
    xvalues3 = Reference(ws, min_col=1, min_row=start_row3+2, max_row=row3-1)
    yvalues3 = Reference(ws, min_col=2, min_row=start_row3+1, max_row=row3-1)
    
    series3 = Series(yvalues3, xvalues3, title_from_data=True)
    chart3.series.append(series3)
    chart3.height = 10
    chart3.width = 15
    
    ws.add_chart(chart3, f'D{start_row3}')
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    
    # Сохранение файла
    filename = 'labs/Lab7_Part2_Graphs.xlsx'
    wb.save(filename)
    print(f"✓ Лабораторная работа №7 (Часть 2) успешно создана: {filename}")
    print(f"  Построено 3 графика функций:")
    print(f"    - y = sin(x)")
    print(f"    - Кусочная функция")
    print(f"    - y = x² - 4")
    return filename

if __name__ == "__main__":
    create_function_graphs()
